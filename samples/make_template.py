"""
Generates 'Master Panel - TEMPLATE.xlsx' — a blank reference copy of the
mapping file that Step 4 produces, so users can see the exact column layout.

Run:  python samples/make_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

OUT = Path(__file__).resolve().parent / "Master Panel - TEMPLATE.xlsx"

BLUE = PatternFill(fill_type="solid", fgColor="074F6A")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = BLUE
        cell.font = WHITE_BOLD
        cell.alignment = CENTER


wb = Workbook()

# Tab 1: Publication Master
ws1 = wb.active
ws1.title = "Publication Master"
ws1.append(["Publication Name", "Publication Identifier"])
ws1.append(["CIO Quarterly Review", 1])
ws1.append(["China Bi-Weekly CN", 2])
ws1.append(["China Bi-Weekly EN", 3])
style_header(ws1, 2)
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 22

# Tab 2: Campaign Mapping
ws2 = wb.create_sheet("Campaign Mapping")
ws2.append([
    "Email Campaign ID",
    "Subject",
    "Sent At (Your time zone)",
    "Identifier",
    "MUTE OR NOT",
    "Into All Publications or not?",
])
# example rows (delete these in the real file – Step 4 fills them automatically)
ws2.append(["123456789", "Example subject line", "2026-01-15", 1, 0, 1])
ws2.append(["987654321", "Another campaign", "2026-01-22", 2, 0, 1])
ws2.append(["555555555", "Muted example", "2026-01-29", 3, 1, 0])
style_header(ws2, 6)
widths = [20, 40, 24, 12, 14, 28]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

wb.save(OUT)
print(f"Wrote {OUT}")

# ============================================
# 04 Hubspot Email List - By Publication
# FULL REFRESH VERSION - rebuild everything every time.
#
# Scans the "04 Hubspot Email List - By Publication" folder tree
# and produces:
#   Output/04/SG Hubspot Email List - BY Publication.xlsx
#   Output/04/Master Panel SG Hubspot Email - BY Publication.xlsx
# ============================================

import os
from pathlib import Path
from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def run(base_dir: Path):
    base_dir = Path(base_dir)

    # =========================
    # 1) Config
    # =========================
    ROOT_DIR = base_dir / "04 Hubspot Email List - By Publication"
    OUTPUT_DIR = base_dir / "Output" / "04"

    BY_PUBLICATION_PATH = OUTPUT_DIR / "SG Hubspot Email List - BY Publication.xlsx"
    MASTER_PANEL_PATH = OUTPUT_DIR / "Master Panel SG Hubspot Email - BY Publication.xlsx"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not ROOT_DIR.exists():
        raise FileNotFoundError(f"[ERROR] Folder not found: {ROOT_DIR}")

    EXCEL_EXTS = {".xlsx", ".xlsm", ".xls"}

    # Stats
    files_scanned = 0
    files_used = 0
    files_skipped = 0
    rows_appended = 0
    publication_count = 0

    # =========================
    # 2) Helper functions
    # =========================
    def collapse_spaces(s: str) -> str:
        while "  " in s:
            s = s.replace("  ", " ")
        return s

    def clean_text(v) -> str:
        if v is None:
            return ""
        s = str(v)
        s = s.replace("\r", " ")
        s = s.replace("\n", " ")
        s = s.replace(chr(160), " ")
        s = collapse_spaces(s).strip()
        return s

    def normalize_bool(v):
        s = clean_text(v).lower()
        if s == "true":
            return "TRUE"
        if s == "false":
            return "FALSE"
        return v

    def normalize_campaign_id(v) -> str:
        return clean_text(v)

    def pad2(n) -> str:
        n = str(n)
        return "0" + n if len(n) == 1 else n

    def extract_date_only(v) -> str:
        s = clean_text(v)
        if s == "":
            return ""
        try:
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                return f"{v.year}-{pad2(v.month)}-{pad2(v.day)}"
        except Exception:
            pass
        if len(s) >= 10:
            return s[:10]
        return s

    def get_last_row(ws, col_num: int) -> int:
        for r in range(ws.max_row, 0, -1):
            val = ws.cell(r, col_num).value
            if clean_text(val) != "":
                return r
        return 1

    def autofit_approx(ws, start_col: int, end_col: int):
        for col in range(start_col, end_col + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row, col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    def should_skip_file(path_obj: Path) -> bool:
        name_lower = path_obj.name.lower()
        if name_lower.startswith("~$"):
            return True
        if path_obj.suffix.lower() not in EXCEL_EXTS:
            return True
        return False

    def infer_publication_name(file_path: Path) -> str:
        rel_parts = file_path.relative_to(ROOT_DIR).parts[:-1]  # exclude filename
        if len(rel_parts) == 0:
            return "Unknown Publication"
        if len(rel_parts) == 1:
            return clean_text(rel_parts[0])
        first = clean_text(rel_parts[0])
        second = clean_text(rel_parts[1]).upper()
        if second in {"CN", "EN"}:
            return f"{first} {second}"
        return first

    # =========================
    # 3) Collect all source files
    # =========================
    all_excel_files = []
    for dirpath, dirnames, filenames in os.walk(ROOT_DIR):
        for fname in filenames:
            full_path = Path(dirpath) / fname
            if should_skip_file(full_path):
                continue
            all_excel_files.append(full_path)

    all_excel_files = sorted(all_excel_files, key=lambda p: str(p).lower())

    if not all_excel_files:
        raise FileNotFoundError(f"[ERROR] No Excel files found under: {ROOT_DIR}")

    # =========================
    # 4) Build publication list
    # =========================
    publication_names = []
    publication_seen = set()

    for f in all_excel_files:
        pub_name = infer_publication_name(f)
        key = pub_name.lower()
        if key not in publication_seen:
            publication_seen.add(key)
            publication_names.append(pub_name)

    publication_names = sorted(publication_names, key=lambda x: x.lower())
    publication_to_id = OrderedDict((pub.lower(), i + 1) for i, pub in enumerate(publication_names))
    publication_count = len(publication_names)

    # =========================
    # 5) Rebuild SG Hubspot Email List - BY Publication.xlsx
    # =========================
    by_pub_wb = Workbook()
    by_pub_ws = by_pub_wb.active
    by_pub_ws.title = "BY Publication"

    headers = [
        "Recipient", "Hub ID", "Email Campaign ID", "SubType", "Subject",
        "Sent At (Your time zone)", "Not Sent Reason", "Not Sent Message",
        "Bounce Reason", "Bounce Message", "Sent", "Delivered", "Suppressed",
        "Dropped", "Bounce", "Spam Report", "Opened", "Clicked", "Unsubscribed",
    ]

    for i, h in enumerate(headers, start=1):
        by_pub_ws.cell(1, i).value = h

    out_row = 2
    seen_campaign_ids = OrderedDict()
    campaign_info = OrderedDict()

    for fil in all_excel_files:
        files_scanned += 1
        print(f"[INFO] Checking: {fil}")

        if fil.suffix.lower() == ".xls":
            print(f"[WARN] Skipped .xls file: {fil}")
            files_skipped += 1
            continue

        try:
            wb = load_workbook(fil, data_only=False, read_only=False)
            ws = wb.worksheets[0]
        except Exception:
            print(f"[WARN] Skipped file (cannot open): {fil}")
            files_skipped += 1
            continue

        last_row_src = get_last_row(ws, 1)
        pub_name = infer_publication_name(fil)
        pub_key = pub_name.lower()
        file_has_new_campaign = False

        file_ids = OrderedDict()
        if last_row_src >= 2:
            for r in range(2, last_row_src + 1):
                recipient = clean_text(ws.cell(r, 1).value)
                cid = normalize_campaign_id(ws.cell(r, 3).value)
                if recipient != "" and cid != "":
                    key = cid.lower()
                    if key not in file_ids:
                        file_ids[key] = cid

        new_ids = OrderedDict()
        for key, cid in file_ids.items():
            if key not in seen_campaign_ids:
                new_ids[key] = cid

        if len(new_ids) == 0:
            print(f"[INFO] Skip (no new Email Campaign ID): {fil}")
            wb.close()
            continue

        if len(file_ids) == 1 and len(new_ids) == 1 and last_row_src >= 2:
            for r in range(2, last_row_src + 1):
                for c in range(1, 20):
                    val = ws.cell(r, c).value
                    by_pub_ws.cell(out_row, c).value = normalize_bool(val)
                out_row += 1
                rows_appended += 1

            for key, cid in new_ids.items():
                seen_campaign_ids[key] = cid
                if key not in campaign_info:
                    campaign_info[key] = {
                        "campaign_id": cid,
                        "subject": clean_text(ws.cell(2, 5).value),
                        "sent_at": extract_date_only(ws.cell(2, 6).value),
                        "publication_name": pub_name,
                        "publication_id": publication_to_id[pub_key],
                    }
            file_has_new_campaign = True

        else:
            for r in range(2, last_row_src + 1):
                recipient = clean_text(ws.cell(r, 1).value)
                cid = normalize_campaign_id(ws.cell(r, 3).value)
                if recipient != "" and cid != "":
                    key = cid.lower()
                    if key in new_ids:
                        for c in range(1, 20):
                            val = ws.cell(r, c).value
                            by_pub_ws.cell(out_row, c).value = normalize_bool(val)
                        out_row += 1
                        rows_appended += 1
                        file_has_new_campaign = True

                        if key not in campaign_info:
                            campaign_info[key] = {
                                "campaign_id": cid,
                                "subject": clean_text(ws.cell(r, 5).value),
                                "sent_at": extract_date_only(ws.cell(r, 6).value),
                                "publication_name": pub_name,
                                "publication_id": publication_to_id[pub_key],
                            }

            for key, cid in new_ids.items():
                seen_campaign_ids[key] = cid

        if file_has_new_campaign:
            files_used += 1
            print(f"[INFO] Added data from: {fil}")

        wb.close()

    campaign_count = len(campaign_info)

    autofit_approx(by_pub_ws, 1, 19)
    by_pub_wb.save(BY_PUBLICATION_PATH)
    by_pub_wb.close()

    print(f"[OK] Rebuilt: {BY_PUBLICATION_PATH}")

    # =========================
    # 6) Rebuild Master Panel from scratch
    # =========================
    master_wb = Workbook()

    ws_master = master_wb.active
    ws_master.title = "Publication Master"
    ws_master.cell(1, 1).value = "Publication Name"
    ws_master.cell(1, 2).value = "Publication Identifier"

    row_master = 2
    for pub_name in publication_names:
        ws_master.cell(row_master, 1).value = pub_name
        ws_master.cell(row_master, 2).value = publication_to_id[pub_name.lower()]
        row_master += 1

    ws_map = master_wb.create_sheet(title="Campaign Mapping")
    ws_map.cell(1, 1).value = "Email Campaign ID"
    ws_map.cell(1, 2).value = "Subject"
    ws_map.cell(1, 3).value = "Sent At (Your time zone)"
    ws_map.cell(1, 4).value = "Identifier"
    ws_map.cell(1, 5).value = "MUTE OR NOT"
    ws_map.cell(1, 6).value = "Into All Publications or not?"

    row_map = 2
    for key, info in campaign_info.items():
        ws_map.cell(row_map, 1).value = info["campaign_id"]
        ws_map.cell(row_map, 2).value = info["subject"]
        ws_map.cell(row_map, 3).value = info["sent_at"]
        ws_map.cell(row_map, 4).value = info["publication_id"]
        ws_map.cell(row_map, 5).value = 0

        pub_name_lower = clean_text(info["publication_name"]).lower()
        if pub_name_lower in {"wealthly daily en", "wealthly daily cn"}:
            ws_map.cell(row_map, 6).value = 0
        else:
            ws_map.cell(row_map, 6).value = 1

        row_map += 1

    autofit_approx(ws_master, 1, 2)
    autofit_approx(ws_map, 1, 6)

    master_wb.save(MASTER_PANEL_PATH)
    master_wb.close()

    print(f"[OK] Rebuilt: {MASTER_PANEL_PATH}")

    # =========================
    # 7) Summary
    # =========================
    print(f"[SUMMARY] Publications found: {publication_count}")
    print(f"[SUMMARY] Campaign IDs found: {campaign_count}")
    print(f"[SUMMARY] Files scanned: {files_scanned}")
    print(f"[SUMMARY] Files used: {files_used}")
    print(f"[SUMMARY] Files skipped: {files_skipped}")
    print(f"[SUMMARY] Rows appended to BY Publication: {rows_appended}")
    print("")
    print("[NEXT] Open the Master Panel file in Output/04 and review the")
    print("       'MUTE OR NOT' and 'Into All Publications or not?' columns")
    print("       in the 'Campaign Mapping' tab BEFORE running Step 5.")

# ============================================
# 05 Hubspot Email list - Masterlist
#
# Reads:
# - Output/03/SG Hubspot Email List - Main.xlsx
# - Output/04/SG Hubspot Email List - BY Publication.xlsx
# - Output/04/Master Panel SG Hubspot Email - BY Publication.xlsx
#
# Output:
# - 05 Hubspot Email list - Masterlist/masterlist.xlsx
#
# Special rules:
# 1) "0 All Publications" only includes campaigns where:
#       - MUTE OR NOT is blank or 0
#       - Into All Publications or not? is blank or 1
# 2) "2. Sales & Client Loyalty" top table only shows Singapore / Hong Kong / Malaysia
# 3) For "Into All Publications or not?" = 0, the campaign is excluded from
#    the All-Publications / loyalty top / internal-external tables, but still
#    counted in the per-publication lower table.
# ============================================

from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from datetime import datetime, date
import re


def run(base_dir: Path):
    base_dir = Path(base_dir)

    # =========================
    # Paths
    # =========================
    panel_path = base_dir / "Output" / "04" / "Master Panel SG Hubspot Email - BY Publication.xlsx"
    main_path = base_dir / "Output" / "03" / "SG Hubspot Email List - Main.xlsx"
    pub_path = base_dir / "Output" / "04" / "SG Hubspot Email List - BY Publication.xlsx"

    output_dir = base_dir / "05 Hubspot Email list - Masterlist"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "masterlist.xlsx"

    missing = []
    if not panel_path.exists():
        missing.append(str(panel_path))
    if not main_path.exists():
        missing.append(str(main_path))
    if not pub_path.exists():
        missing.append(str(pub_path))

    if missing:
        raise FileNotFoundError(
            "Missing required file(s) (run Steps 3 and 4 first):\n- " + "\n- ".join(missing)
        )

    # =========================
    # Constants / styles
    # =========================
    SUMMARY_SHEET_NAME = "1.1 Email Open Click by Country"
    ANALYSIS_SHEET_NAME = "2. Sales & Client Loyalty"
    ALL_SHEET_NAME = "0 All Publications"

    DISPLAY_COUNTRIES_SUMMARY = ["Singapore", "Hong Kong", "Malaysia", "Thailand", "Others"]
    DISPLAY_COUNTRIES_ANALYSIS = ["Singapore", "Hong Kong", "Malaysia"]  # only top table
    DISPLAY_COUNTRIES_CONTACT = ["Singapore", "Hong Kong", "Malaysia", "Thailand", "Others"]

    BLUE_FILL = PatternFill(fill_type="solid", fgColor="074F6A")
    WHITE_FONT_BOLD = Font(color="FFFFFF", bold=True)
    BOLD_FONT = Font(bold=True)

    LIGHT_GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
    DARK_GREEN_FONT = Font(color="006100")
    LIGHT_ORANGE_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
    DARK_ORANGE_FONT = Font(color="9C5700")
    LIGHT_RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
    DARK_RED_FONT = Font(color="9C0006")

    THIN_BLACK_SIDE = Side(style="thin", color="000000")
    BORDER_ALL = Border(
        left=THIN_BLACK_SIDE, right=THIN_BLACK_SIDE,
        top=THIN_BLACK_SIDE, bottom=THIN_BLACK_SIDE,
    )

    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

    # =========================
    # Utility functions
    # =========================
    def collapse_spaces(s):
        return re.sub(r"\s+", " ", s).strip()

    def clean_text(v):
        if v is None:
            return ""
        s = str(v).replace("\r", " ").replace("\n", " ").replace(chr(160), " ")
        return collapse_spaces(s)

    def normalize_email(v):
        return clean_text(v).lower()

    def is_true_value(v):
        if isinstance(v, bool):
            return v
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v) == 1
        s = clean_text(v).upper()
        return s in {"TRUE", "1", "YES"}

    def is_included_mute(v):
        if v is None or v == "":
            return True
        if isinstance(v, bool):
            return v is False
        if isinstance(v, (int, float)):
            return float(v) == 0
        s = clean_text(v).upper()
        return s in {"", "0", "0.0", "FALSE", "NO"}

    def is_into_all_publications(v):
        if v is None or v == "":
            return True
        if isinstance(v, bool):
            return v is True
        if isinstance(v, (int, float)):
            return float(v) == 1
        s = clean_text(v).upper()
        return s in {"", "1", "1.0", "TRUE", "YES"}

    def round2(v):
        return round(float(v), 2)

    def summary_country(v):
        s = clean_text(v)
        return s if s else "(Blank)"

    def country_key(v):
        return clean_text(v).lower().replace(" ", "").replace("-", "")

    def map_country_to_display(v):
        key = country_key(v)
        if key == "singapore":
            return "Singapore"
        if key == "hongkong":
            return "Hong Kong"
        if key == "malaysia":
            return "Malaysia"
        if key == "thailand":
            return "Thailand"
        return "Others"

    def engagement_tier(rate_pct):
        if rate_pct >= 50:
            return "Loyal"
        if rate_pct >= 25:
            return "Casual"
        return "Dormant"

    def normalize_tier_contact_type(v):
        s = clean_text(v).lower().replace("-", " ")
        s = collapse_spaces(s)
        if s == "tier 1":
            return "Tier 1"
        if s == "tier 2":
            return "Tier 2"
        if s == "tier 3":
            return "Tier 3"
        return ""

    def tier_index_from_label(tier_label):
        s = clean_text(tier_label).lower()
        if s == "tier 1":
            return 0
        if s == "tier 2":
            return 1
        if s == "tier 3":
            return 2
        return -1

    def normalize_contact_bucket(v):
        s = clean_text(v).lower().replace("-", " ")
        s = collapse_spaces(s)
        if s in {"internal", "external"}:
            return s.capitalize()
        return ""

    def make_state_key(campaign_id, email_val):
        return f"{campaign_id}||{str(email_val).lower()}"

    def state_rank(v):
        v = str(v)
        if v == "NR":
            return 0
        if v == "0":
            return 1
        if v == "1":
            return 2
        if v == "2":
            return 3
        return -1

    def pick_better_state(old_state, new_state):
        return new_state if state_rank(new_state) > state_rank(old_state) else old_state

    def determine_state(sent_v, delivered_v, opened_v, clicked_v):
        sent_b = is_true_value(sent_v)
        delivered_b = is_true_value(delivered_v)
        opened_b = is_true_value(opened_v)
        clicked_b = is_true_value(clicked_v)

        if (not sent_b) or (not delivered_b):
            return "NR"
        if (opened_b is False) and (clicked_b is False):
            return "0"
        if (opened_b is True) and (clicked_b is False):
            return "1"
        if (opened_b is True) and (clicked_b is True):
            return "2"
        return "0"

    def parse_possible_excel_date(v):
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            try:
                return from_excel(v).date()
            except Exception:
                return None

        s = clean_text(v)
        if not s:
            return None

        s2 = s.replace("/", "-")
        if len(s2) >= 10 and s2[:4].isdigit() and s2[4] == "-" and s2[7] == "-":
            try:
                return datetime.strptime(s2[:10], "%Y-%m-%d").date()
            except Exception:
                pass

        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue

        return None

    def build_display_date(v):
        dt = parse_possible_excel_date(v)
        if dt is not None:
            return f"{dt.day}/{dt.month}/{dt.year}"
        return clean_text(v)

    def get_date_serial_value(v):
        dt = parse_possible_excel_date(v)
        if dt is not None:
            return dt.toordinal()
        return 0

    def make_safe_sheet_name(s):
        t = clean_text(s)
        for ch in ['\\', '/', ':', '*', '?', '[', ']']:
            t = t.replace(ch, ' ')
        t = collapse_spaces(t)
        if not t:
            t = "Sheet"
        return t[:31]

    def unique_sheet_name(wb, base_name):
        clean_base = make_safe_sheet_name(base_name)
        candidate = clean_base
        idx = 1
        existing = {ws.title.lower() for ws in wb.worksheets}
        while candidate.lower() in existing:
            suffix = f" {idx}"
            candidate = (clean_base[:31 - len(suffix)] + suffix).strip()
            idx += 1
        return candidate

    def apply_simple_borders(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = BORDER_ALL

    def apply_blue_header(ws, cell_range):
        for row in ws[cell_range]:
            for cell in row:
                cell.fill = BLUE_FILL
                cell.font = WHITE_FONT_BOLD
                cell.border = BORDER_ALL
                cell.alignment = CENTER_ALIGN

    def apply_total_row_blue(ws, row_num, start_col, end_col):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = BLUE_FILL
            cell.font = WHITE_FONT_BOLD
            cell.border = BORDER_ALL
            if c == 1 or c == 3:
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN

    def add_hyperlink(ws, row_num, col_num, target_sheet, display_text):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = display_text
        cell.hyperlink = f"#'{target_sheet}'!A1"
        cell.style = "Hyperlink"
        cell.alignment = CENTER_ALIGN

    def best_fit_columns(ws, min_width=8, max_width=40):
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

    def set_number_format_range(ws, start_row, end_row, start_col, end_col, number_format):
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

    def apply_open_rate_style(cell):
        val = cell.value
        if val in (None, "", "NA") or not isinstance(val, (int, float)):
            return
        if val > 20:
            cell.fill = LIGHT_GREEN_FILL
            cell.font = DARK_GREEN_FONT
        elif 10 < val < 20:
            cell.fill = LIGHT_ORANGE_FILL
            cell.font = DARK_ORANGE_FONT
        elif val < 10:
            cell.fill = LIGHT_RED_FILL
            cell.font = DARK_RED_FONT

    def apply_click_rate_style(cell):
        val = cell.value
        if val in (None, "", "NA") or not isinstance(val, (int, float)):
            return
        if val > 5:
            cell.fill = LIGHT_GREEN_FILL
            cell.font = DARK_GREEN_FONT
        elif 2 < val < 5:
            cell.fill = LIGHT_ORANGE_FILL
            cell.font = DARK_ORANGE_FONT
        elif val < 2:
            cell.fill = LIGHT_RED_FILL
            cell.font = DARK_RED_FONT

    def apply_summary_conditional_format(ws, start_row, end_row, skip_blue_rows=None):
        skip_blue_rows = set(skip_blue_rows or [])
        for r in range(start_row, end_row + 1):
            if r in skip_blue_rows:
                continue
            apply_open_rate_style(ws.cell(r, 5))
            apply_click_rate_style(ws.cell(r, 6))

    # =========================
    # Read input workbooks
    # =========================
    dict_pub_names = {}
    pub_ids = []
    pub_names = []

    dict_relevant_campaign = {}
    campaigns = []

    dict_all_campaign = {}
    all_campaigns_unfiltered = []

    dict_main_emails = {}
    main_rows = []

    dict_state = {}

    print("[INFO] Reading Master Panel...")
    wb_panel = load_workbook(panel_path, data_only=True, read_only=True)
    if "Publication Master" not in wb_panel.sheetnames:
        raise ValueError("Sheet not found in Master Panel: Publication Master")
    if "Campaign Mapping" not in wb_panel.sheetnames:
        raise ValueError("Sheet not found in Master Panel: Campaign Mapping")

    ws_pub_master = wb_panel["Publication Master"]
    for row in ws_pub_master.iter_rows(min_row=2, values_only=True):
        pub_name = clean_text(row[0] if len(row) > 0 else None)
        ident = clean_text(row[1] if len(row) > 1 else None)
        if ident and ident not in dict_pub_names:
            dict_pub_names[ident] = pub_name
            pub_ids.append(ident)
            pub_names.append(pub_name)

    if not pub_ids:
        raise ValueError("No Publication Identifier found in Publication Master.")

    ws_camp_map = wb_panel["Campaign Mapping"]
    for row in ws_camp_map.iter_rows(min_row=2, values_only=True):
        cid = clean_text(row[0] if len(row) > 0 else None)
        sent_date_raw = row[2] if len(row) > 2 else None
        ident = clean_text(row[3] if len(row) > 3 else None)
        mute_val = row[4] if len(row) > 4 else None
        into_all_val = row[5] if len(row) > 5 else None

        if cid and ident and ident in dict_pub_names and is_included_mute(mute_val):
            if cid not in dict_all_campaign:
                dict_all_campaign[cid] = True
                all_campaigns_unfiltered.append({
                    "id": cid,
                    "pub_id": ident,
                    "date_text": build_display_date(sent_date_raw),
                    "date_serial": get_date_serial_value(sent_date_raw),
                    "into_all": is_into_all_publications(into_all_val),
                })

            if is_into_all_publications(into_all_val):
                if cid not in dict_relevant_campaign:
                    dict_relevant_campaign[cid] = True
                    campaigns.append({
                        "id": cid,
                        "pub_id": ident,
                        "date_text": build_display_date(sent_date_raw),
                        "date_serial": get_date_serial_value(sent_date_raw),
                        "into_all": True,
                    })
    wb_panel.close()

    print("[INFO] Reading Main...")
    wb_main = load_workbook(main_path, data_only=True, read_only=True)
    ws_main = wb_main.worksheets[0]
    for row in ws_main.iter_rows(min_row=2, values_only=True):
        email_val = normalize_email(row[0] if len(row) > 0 else None)
        name_val = clean_text(row[1] if len(row) > 1 else None)
        country_val = clean_text(row[2] if len(row) > 2 else None)
        ctype_val = clean_text(row[3] if len(row) > 3 else None)

        if email_val and email_val not in dict_main_emails:
            dict_main_emails[email_val] = True
            main_rows.append({
                "email": email_val,
                "name": name_val,
                "country": country_val,
                "contact_type": ctype_val,
            })
    wb_main.close()

    if not main_rows:
        raise ValueError("No valid rows found in SG Hubspot Email List - Main.xlsx")

    print("[INFO] Reading BY Publication...")
    wb_pub = load_workbook(pub_path, data_only=True, read_only=True)
    ws_pub = wb_pub.worksheets[0]
    for row in ws_pub.iter_rows(min_row=2, values_only=True):
        recipient = normalize_email(row[0] if len(row) > 0 else None)
        cid = clean_text(row[2] if len(row) > 2 else None)

        if recipient and cid and cid in dict_all_campaign and recipient in dict_main_emails:
            sent_val = row[10] if len(row) > 10 else None
            delivered_val = row[11] if len(row) > 11 else None
            opened_val = row[16] if len(row) > 16 else None
            clicked_val = row[17] if len(row) > 17 else None

            state_val = determine_state(sent_val, delivered_val, opened_val, clicked_val)
            key_val = make_state_key(cid, recipient)

            if key_val in dict_state:
                dict_state[key_val] = pick_better_state(dict_state[key_val], state_val)
            else:
                dict_state[key_val] = state_val
    wb_pub.close()

    # =========================
    # Campaign helpers
    # =========================
    def compare_campaign_tuple(c):
        return (float(c["date_serial"]), str(c["id"]).lower())

    def get_campaigns_for_publication(pub_id):
        local = [c.copy() for c in all_campaigns_unfiltered if str(c["pub_id"]).lower() == str(pub_id).lower()]
        local.sort(key=compare_campaign_tuple)
        return local

    def get_campaigns_for_publication_unfiltered(pub_id):
        local = [c.copy() for c in all_campaigns_unfiltered if str(c["pub_id"]).lower() == str(pub_id).lower()]
        local.sort(key=compare_campaign_tuple)
        return local

    def get_all_campaigns():
        local = [c.copy() for c in campaigns]
        local.sort(key=compare_campaign_tuple)
        return local

    def get_engagement_tier_for_email(email_now, local_campaigns):
        valid_cnt = 0
        open_cnt = 0
        for c in local_campaigns:
            state_now = dict_state.get(make_state_key(c["id"], email_now), "NA")
            if state_now == "0":
                valid_cnt += 1
            elif state_now in {"1", "2"}:
                valid_cnt += 1
                open_cnt += 1
        if valid_cnt == 0:
            return "NA"
        return engagement_tier((open_cnt / valid_cnt) * 100)

    summary_row = 2
    summary_total_rows = []

    # =========================
    # Sheet renderers
    # =========================
    def render_campaign_tab(ws, tab_name, local_campaigns, add_to_summary, ws_summary):
        nonlocal summary_row, summary_total_rows

        camp_n = len(local_campaigns)
        total_arr = [0] * camp_n
        open_arr = [0] * camp_n
        click_arr = [0] * camp_n

        country_valid = {k: 0 for k in DISPLAY_COUNTRIES_SUMMARY}
        country_open = {k: 0 for k in DISPLAY_COUNTRIES_SUMMARY}
        country_click = {k: 0 for k in DISPLAY_COUNTRIES_SUMMARY}

        overall_valid = 0
        overall_open = 0
        overall_click = 0
        latest_total_valid = 0
        latest_total_open = 0
        latest_total_click = 0
        latest_idx = camp_n - 1

        add_hyperlink(ws, 1, 1, SUMMARY_SHEET_NAME, "Back to 1.1")
        ws.cell(2, 1).value = "Avg Open Rate (%)"
        ws.cell(3, 1).value = "Avg Click Rate (%)"

        ws.cell(1, 6).value = "Email Campaign ID"
        ws.cell(2, 6).value = "Daily Open Rate (%)"
        ws.cell(3, 6).value = "Daily Click Rate (%)"
        ws.cell(4, 6).value = "Total Recipient"

        headers = ["Email", "Name", "Country", "Contact Type", "Engagement Tier", "Open Rate (%)"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(5, idx).value = header

        for j, c in enumerate(local_campaigns):
            col_num = 7 + j
            ws.cell(1, col_num).value = c["id"]
            ws.cell(5, col_num).value = c["date_text"]

        for i, row_data in enumerate(main_rows):
            row_num = 6 + i
            email_now = row_data["email"]
            country_now = summary_country(row_data["country"])

            ws.cell(row_num, 1).value = row_data["email"]
            ws.cell(row_num, 2).value = row_data["name"]
            ws.cell(row_num, 3).value = row_data["country"]
            ws.cell(row_num, 4).value = row_data["contact_type"]

            valid_cnt = 0
            open_cnt = 0
            click_cnt = 0

            for j, c in enumerate(local_campaigns):
                col_num = 7 + j
                state_now = dict_state.get(make_state_key(c["id"], email_now), "NA")
                ws.cell(row_num, col_num).value = state_now

                if state_now == "0":
                    valid_cnt += 1
                    total_arr[j] += 1
                elif state_now == "1":
                    valid_cnt += 1
                    open_cnt += 1
                    total_arr[j] += 1
                    open_arr[j] += 1
                elif state_now == "2":
                    valid_cnt += 1
                    open_cnt += 1
                    click_cnt += 1
                    total_arr[j] += 1
                    open_arr[j] += 1
                    click_arr[j] += 1

            if valid_cnt == 0:
                ws.cell(row_num, 5).value = "NA"
                ws.cell(row_num, 6).value = "NA"
            else:
                rate_val = round2((open_cnt / valid_cnt) * 100)
                ws.cell(row_num, 6).value = rate_val
                ws.cell(row_num, 5).value = engagement_tier(rate_val)

            overall_valid += valid_cnt
            overall_open += open_cnt
            overall_click += click_cnt

            if camp_n > 0:
                latest_state = dict_state.get(make_state_key(local_campaigns[latest_idx]["id"], email_now), "NA")
                mapped_country = map_country_to_display(country_now)

                if latest_state == "0":
                    country_valid[mapped_country] += 1
                    latest_total_valid += 1
                elif latest_state == "1":
                    country_valid[mapped_country] += 1
                    country_open[mapped_country] += 1
                    latest_total_valid += 1
                    latest_total_open += 1
                elif latest_state == "2":
                    country_valid[mapped_country] += 1
                    country_open[mapped_country] += 1
                    country_click[mapped_country] += 1
                    latest_total_valid += 1
                    latest_total_open += 1
                    latest_total_click += 1

        if overall_valid == 0:
            ws.cell(2, 2).value = "NA"
            ws.cell(3, 2).value = "NA"
        else:
            ws.cell(2, 2).value = round2((overall_open / overall_valid) * 100)
            ws.cell(3, 2).value = round2((overall_click / overall_valid) * 100)

        for j in range(camp_n):
            col_num = 7 + j
            ws.cell(4, col_num).value = total_arr[j]
            if total_arr[j] == 0:
                ws.cell(2, col_num).value = "NA"
                ws.cell(3, col_num).value = "NA"
            else:
                ws.cell(2, col_num).value = round2((open_arr[j] / total_arr[j]) * 100)
                ws.cell(3, col_num).value = round2((click_arr[j] / total_arr[j]) * 100)

        if add_to_summary:
            group_start_row = summary_row

            for country_name in DISPLAY_COUNTRIES_SUMMARY:
                ws_summary.cell(summary_row, 3).value = country_name
                ws_summary.cell(summary_row, 4).value = country_valid[country_name]
                if country_valid[country_name] == 0:
                    ws_summary.cell(summary_row, 5).value = "NA"
                    ws_summary.cell(summary_row, 6).value = "NA"
                else:
                    ws_summary.cell(summary_row, 5).value = round2((country_open[country_name] / country_valid[country_name]) * 100)
                    ws_summary.cell(summary_row, 6).value = round2((country_click[country_name] / country_valid[country_name]) * 100)
                summary_row += 1

            ws_summary.cell(summary_row, 3).value = "Total"
            ws_summary.cell(summary_row, 4).value = latest_total_valid
            if latest_total_valid == 0:
                ws_summary.cell(summary_row, 5).value = "NA"
                ws_summary.cell(summary_row, 6).value = "NA"
            else:
                ws_summary.cell(summary_row, 5).value = round2((latest_total_open / latest_total_valid) * 100)
                ws_summary.cell(summary_row, 6).value = round2((latest_total_click / latest_total_valid) * 100)

            total_row = summary_row
            summary_total_rows.append(total_row)

            ws_summary.cell(group_start_row, 1).value = tab_name
            ws_summary.cell(group_start_row, 2).value = camp_n

            if total_row > group_start_row:
                ws_summary.merge_cells(start_row=group_start_row, start_column=1, end_row=total_row, end_column=1)
                ws_summary.merge_cells(start_row=group_start_row, start_column=2, end_row=total_row, end_column=2)

            add_hyperlink(ws_summary, group_start_row, 1, tab_name, tab_name)

            for r in range(group_start_row, total_row + 1):
                ws_summary.cell(r, 1).alignment = CENTER_ALIGN
                ws_summary.cell(r, 2).alignment = CENTER_ALIGN
                ws_summary.cell(r, 3).alignment = LEFT_ALIGN
                ws_summary.cell(r, 4).alignment = CENTER_ALIGN
                ws_summary.cell(r, 5).alignment = RIGHT_ALIGN
                ws_summary.cell(r, 6).alignment = RIGHT_ALIGN

            apply_simple_borders(ws_summary, f"A{group_start_row}:F{total_row}")
            apply_total_row_blue(ws_summary, total_row, 3, 6)
            summary_row += 1

        last_col = max(6 + camp_n, 6)
        last_row_sheet = 5 + len(main_rows)

        apply_blue_header(ws, f"A5:{get_column_letter(last_col)}5")
        apply_blue_header(ws, f"F1:{get_column_letter(last_col)}1")
        apply_blue_header(ws, "F2:F4")
        apply_simple_borders(ws, f"A1:{get_column_letter(last_col)}{last_row_sheet}")

        ws["B2"].number_format = "0.00"
        ws["B3"].number_format = "0.00"
        set_number_format_range(ws, 6, last_row_sheet, 6, 6, "0.00")
        if camp_n > 0:
            set_number_format_range(ws, 2, 3, 7, last_col, "0.00")

        ws.auto_filter.ref = f"A5:{get_column_letter(last_col)}{last_row_sheet}"
        best_fit_columns(ws)

    def render_simple_contacttype_table(ws, start_row, contact_type_label, counts_dict):
        header_row = start_row
        data_start_row = start_row + 1

        ws.cell(header_row, 1).value = "Country"
        ws.cell(header_row, 2).value = "Contact Type"
        ws.cell(header_row, 3).value = "Loyal"
        ws.cell(header_row, 4).value = "Casual"
        ws.cell(header_row, 5).value = "Dormant"
        apply_blue_header(ws, f"A{header_row}:E{header_row}")

        grand_loyal = 0
        grand_casual = 0
        grand_dormant = 0

        for i, country_name in enumerate(DISPLAY_COUNTRIES_CONTACT):
            row_num = data_start_row + i
            loyal_cnt = counts_dict.get(country_name, {}).get("Loyal", 0)
            casual_cnt = counts_dict.get(country_name, {}).get("Casual", 0)
            dormant_cnt = counts_dict.get(country_name, {}).get("Dormant", 0)

            grand_loyal += loyal_cnt
            grand_casual += casual_cnt
            grand_dormant += dormant_cnt

            ws.cell(row_num, 1).value = country_name
            ws.cell(row_num, 2).value = contact_type_label
            ws.cell(row_num, 3).value = loyal_cnt
            ws.cell(row_num, 4).value = casual_cnt
            ws.cell(row_num, 5).value = dormant_cnt

        gt_row = data_start_row + len(DISPLAY_COUNTRIES_CONTACT)
        ws.cell(gt_row, 1).value = "Grand Total"
        ws.cell(gt_row, 2).value = contact_type_label
        ws.cell(gt_row, 3).value = grand_loyal
        ws.cell(gt_row, 4).value = grand_casual
        ws.cell(gt_row, 5).value = grand_dormant

        apply_simple_borders(ws, f"A{header_row}:E{gt_row}")
        apply_total_row_blue(ws, gt_row, 1, 5)

        for r in range(data_start_row, gt_row + 1):
            ws.cell(r, 1).alignment = LEFT_ALIGN
            ws.cell(r, 2).alignment = LEFT_ALIGN
            for c in range(3, 6):
                ws.cell(r, c).alignment = CENTER_ALIGN

    def render_tab3_publication_country_internal_external(ws, start_row):
        publication_tabs = []
        for ident, pub_name in zip(pub_ids, pub_names):
            tab_name = f"{ident} {pub_name}"
            local_campaigns = get_campaigns_for_publication_unfiltered(ident)
            publication_tabs.append({
                "tab_name": tab_name,
                "campaigns": local_campaigns,
            })

        header_row_1 = start_row
        header_row_2 = start_row + 1
        row_ptr = start_row + 2

        ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
        ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
        ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_1, end_column=7)
        ws.merge_cells(start_row=header_row_1, start_column=8, end_row=header_row_2, end_column=8)

        ws.cell(header_row_1, 1).value = "Publication Name"
        ws.cell(header_row_1, 2).value = "Category"
        ws.cell(header_row_1, 3).value = "Country"
        ws.cell(header_row_2, 3).value = "Singapore"
        ws.cell(header_row_2, 4).value = "Hong Kong"
        ws.cell(header_row_2, 5).value = "Malaysia"
        ws.cell(header_row_2, 6).value = "Thailand"
        ws.cell(header_row_2, 7).value = "Others"
        ws.cell(header_row_1, 8).value = "Total Recipient"

        for pub in publication_tabs:
            pub_start_row = row_ptr
            counts_by_bucket = {
                "Internal": {"Singapore": 0, "Hong Kong": 0, "Malaysia": 0, "Thailand": 0, "Others": 0, "Total": 0},
                "External": {"Singapore": 0, "Hong Kong": 0, "Malaysia": 0, "Thailand": 0, "Others": 0, "Total": 0},
            }

            local_campaigns = pub["campaigns"]
            if local_campaigns:
                latest_campaign = local_campaigns[-1]
                latest_campaign_id = latest_campaign["id"]

                for row_data in main_rows:
                    email_now = row_data["email"]
                    bucket_now = normalize_contact_bucket(row_data["contact_type"])
                    if bucket_now not in {"Internal", "External"}:
                        continue

                    state_now = dict_state.get(make_state_key(latest_campaign_id, email_now), "NA")
                    if state_now not in {"0", "1", "2"}:
                        continue

                    mapped_country = map_country_to_display(row_data["country"])
                    counts_by_bucket[bucket_now][mapped_country] += 1
                    counts_by_bucket[bucket_now]["Total"] += 1

            ws.cell(row_ptr, 2).value = "Internal"
            ws.cell(row_ptr, 3).value = counts_by_bucket["Internal"]["Singapore"]
            ws.cell(row_ptr, 4).value = counts_by_bucket["Internal"]["Hong Kong"]
            ws.cell(row_ptr, 5).value = counts_by_bucket["Internal"]["Malaysia"]
            ws.cell(row_ptr, 6).value = counts_by_bucket["Internal"]["Thailand"]
            ws.cell(row_ptr, 7).value = counts_by_bucket["Internal"]["Others"]
            ws.cell(row_ptr, 8).value = counts_by_bucket["Internal"]["Total"]

            ws.cell(row_ptr + 1, 2).value = "External"
            ws.cell(row_ptr + 1, 3).value = counts_by_bucket["External"]["Singapore"]
            ws.cell(row_ptr + 1, 4).value = counts_by_bucket["External"]["Hong Kong"]
            ws.cell(row_ptr + 1, 5).value = counts_by_bucket["External"]["Malaysia"]
            ws.cell(row_ptr + 1, 6).value = counts_by_bucket["External"]["Thailand"]
            ws.cell(row_ptr + 1, 7).value = counts_by_bucket["External"]["Others"]
            ws.cell(row_ptr + 1, 8).value = counts_by_bucket["External"]["Total"]

            ws.cell(pub_start_row, 1).value = pub["tab_name"]
            ws.merge_cells(start_row=pub_start_row, start_column=1, end_row=pub_start_row + 1, end_column=1)
            ws.cell(pub_start_row, 1).alignment = CENTER_ALIGN

            apply_simple_borders(ws, f"A{pub_start_row}:H{pub_start_row + 1}")
            row_ptr += 2

        apply_blue_header(ws, f"A{header_row_1}:H{header_row_2}")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 16

    def render_sales_loyalty_analysis(ws):
        filtered_all_campaigns = get_all_campaigns()

        country_counts = {k: [[0, 0, 0], [0, 0, 0], [0, 0, 0]] for k in DISPLAY_COUNTRIES_ANALYSIS}
        internal_counts = {k: {"Loyal": 0, "Casual": 0, "Dormant": 0} for k in DISPLAY_COUNTRIES_CONTACT}
        external_counts = {k: {"Loyal": 0, "Casual": 0, "Dormant": 0} for k in DISPLAY_COUNTRIES_CONTACT}

        def add_loyalty_count(matrix_obj, tier_label, loyalty_idx):
            tier_idx = tier_index_from_label(tier_label)
            if tier_idx >= 0 and 0 <= loyalty_idx <= 2:
                matrix_obj[tier_idx][loyalty_idx] += 1

        def get_matrix_val(matrix_obj, tier_idx, loyalty_idx):
            return matrix_obj[tier_idx][loyalty_idx]

        ws.cell(1, 1).value = "Country"
        ws.cell(1, 2).value = "Sales Tier"
        ws.cell(1, 3).value = "Loyal"
        ws.cell(1, 4).value = "Casual"
        ws.cell(1, 5).value = "Dormant"
        ws.cell(1, 6).value = "Tier Total"
        apply_blue_header(ws, "A1:F1")

        for row_data in main_rows:
            email_now = row_data["email"]
            mapped_country = map_country_to_display(row_data["country"])
            contact_now = normalize_tier_contact_type(row_data["contact_type"])
            bucket_now = normalize_contact_bucket(row_data["contact_type"])
            engagement_now = get_engagement_tier_for_email(email_now, filtered_all_campaigns)

            if contact_now and mapped_country in country_counts:
                if engagement_now == "Loyal":
                    add_loyalty_count(country_counts[mapped_country], contact_now, 0)
                elif engagement_now == "Casual":
                    add_loyalty_count(country_counts[mapped_country], contact_now, 1)
                elif engagement_now == "Dormant":
                    add_loyalty_count(country_counts[mapped_country], contact_now, 2)

            if bucket_now in {"Internal", "External"} and engagement_now in {"Loyal", "Casual", "Dormant"}:
                target_dict = internal_counts if bucket_now == "Internal" else external_counts
                target_dict[mapped_country][engagement_now] += 1

        row_num = 2
        for country_name in DISPLAY_COUNTRIES_ANALYSIS:
            start_row = row_num
            matrix = country_counts[country_name]

            for tier_idx, tier_label in enumerate(["Tier 1", "Tier 2", "Tier 3"]):
                loyal_cnt = get_matrix_val(matrix, tier_idx, 0)
                casual_cnt = get_matrix_val(matrix, tier_idx, 1)
                dormant_cnt = get_matrix_val(matrix, tier_idx, 2)
                total_cnt = loyal_cnt + casual_cnt + dormant_cnt

                ws.cell(row_num, 2).value = tier_label
                ws.cell(row_num, 3).value = loyal_cnt
                ws.cell(row_num, 4).value = casual_cnt
                ws.cell(row_num, 5).value = dormant_cnt
                ws.cell(row_num, 6).value = total_cnt
                row_num += 1

            loyal_cnt = sum(matrix[tier][0] for tier in range(3))
            casual_cnt = sum(matrix[tier][1] for tier in range(3))
            dormant_cnt = sum(matrix[tier][2] for tier in range(3))
            total_cnt = loyal_cnt + casual_cnt + dormant_cnt

            ws.cell(row_num, 2).value = "Grand Total"
            ws.cell(row_num, 3).value = loyal_cnt
            ws.cell(row_num, 4).value = casual_cnt
            ws.cell(row_num, 5).value = dormant_cnt
            ws.cell(row_num, 6).value = total_cnt

            end_row = row_num
            ws.cell(start_row, 1).value = country_name
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

            for r in range(start_row, end_row + 1):
                ws.cell(r, 1).alignment = CENTER_ALIGN
                ws.cell(r, 2).alignment = LEFT_ALIGN
                for c in range(3, 7):
                    ws.cell(r, c).alignment = CENTER_ALIGN

            apply_simple_borders(ws, f"A{start_row}:F{end_row}")
            apply_total_row_blue(ws, end_row, 2, 6)
            row_num += 1

        internal_start_row = row_num + 2
        render_simple_contacttype_table(ws, internal_start_row, "Internal", internal_counts)

        external_start_row = internal_start_row + len(DISPLAY_COUNTRIES_CONTACT) + 5
        render_simple_contacttype_table(ws, external_start_row, "External", external_counts)

        tab3_start_row = external_start_row + len(DISPLAY_COUNTRIES_CONTACT) + 6
        render_tab3_publication_country_internal_external(ws, tab3_start_row)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 14
        for col in ["C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 12

    def render_all_publications_tab(ws):
        render_campaign_tab(ws, ALL_SHEET_NAME, get_all_campaigns(), False, ws_summary)

    def render_publication_tab(ws, pub_id, tab_name):
        render_campaign_tab(ws, tab_name, get_campaigns_for_publication(pub_id), True, ws_summary)

    # =========================
    # Build workbook
    # =========================
    print("[INFO] Building masterlist...")
    wb_out = Workbook()
    ws_summary = wb_out.active
    ws_summary.title = SUMMARY_SHEET_NAME
    ws_analysis = wb_out.create_sheet(ANALYSIS_SHEET_NAME)
    ws_all = wb_out.create_sheet(ALL_SHEET_NAME)

    ws_summary.cell(1, 1).value = "Publication Name"
    ws_summary.cell(1, 2).value = "Sent Number"
    ws_summary.cell(1, 3).value = "Country"
    ws_summary.cell(1, 4).value = "Total Recipient"
    ws_summary.cell(1, 5).value = "Open Rate (%)"
    ws_summary.cell(1, 6).value = "Click Rate (%)"

    render_all_publications_tab(ws_all)
    render_sales_loyalty_analysis(ws_analysis)

    for ident, pub_name in zip(pub_ids, pub_names):
        desired_name = f"{ident} {pub_name}"
        actual_name = unique_sheet_name(wb_out, desired_name)
        ws_out = wb_out.create_sheet(actual_name)
        render_publication_tab(ws_out, ident, actual_name)

    apply_blue_header(ws_summary, "A1:F1")
    if summary_row > 2:
        set_number_format_range(ws_summary, 2, summary_row - 1, 5, 6, "0.00")
        apply_summary_conditional_format(ws_summary, 2, summary_row - 1, skip_blue_rows=summary_total_rows)

    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 14
    ws_summary.column_dimensions["E"].width = 16
    ws_summary.column_dimensions["F"].width = 16

    wb_out.save(output_path)
    print(f"[OK] Created: {output_path}")
